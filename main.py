from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#load existing file
wb = load_workbook('C:/og/pyxl/Copy of Spend based Utility.xlsx')

#create an active worksheet
ws = wb.active

# grab a column
# column_b = ws['B']

# for cell in column_b:
#     print(cell.value)

#grab a range
range = ws['A2' : 'B22']

for cell in range:
    for x in cell:
        print(x.value)
