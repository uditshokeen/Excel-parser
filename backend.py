from flask import Flask, request, jsonify
from openpyxl import load_workbook
from io import BytesIO

app = Flask(__name__)

# Route to render the upload form (HTML)
@app.route('/')
def upload_form():
    return '''
    <!doctype html>
    <html>
        <head>
            <title>Upload Excel File</title>
        </head>
        <body>
            <h2>Upload an Excel File</h2>
            <form action="/parse-excel" method="POST" enctype="multipart/form-data">
                <input type="file" name="file" accept=".xlsx, .xls">
                <br><br>
                <input type="submit" value="Upload and Parse">
            </form>
        </body>
    </html>
    '''

# Route to handle Excel file upload and processing
@app.route('/parse-excel', methods=['POST'])
def parse_excel():
    try:
        # Check if the file is in the request
        if 'file' not in request.files:
            return jsonify({"error": "No file part in the request"}), 400

        file = request.files['file']

        # Check if a file was selected
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400

        # Read the file 
        file_stream = BytesIO(file.read())
        wb = load_workbook(file_stream)
        ws = wb.active

        # Define the range
        start_row = 2
        end_row = ws.max_row
        start_col = 1
        end_col = ws.max_column

        # Get headers for the selected columns
        headers = [ws.cell(row=1, column=col).value 
                   for col in range(start_col, end_col + 1)]

        result = []

        # Iterate through the specified range
        for idx, row in enumerate(
            ws.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_col,
                max_col=end_col,
                values_only=True
            ),
            start=start_row
        ):
            row_data = {}
            for i in range(len(headers)):
                row_data[headers[i]] = row[i]
            result.append(row_data)

        # Return the parsed data as JSON
        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)}), 400


if __name__ == '__main__':
    app.run(debug=True)
