from openpyxl import load_workbook

wb = load_workbook('C:/og/pyxl/Copy of Spend based Utility.xlsx')

ws = wb.active

# Define the range
start_row = 2   
end_row = ws.max_row  
start_col = 1   
end_col = ws.max_column   

# Get headers for the selected columns
headers = [ws.cell(row=1, column=col).value 
for col in range(start_col, end_col + 1)]

# Iterate through the specified range
for idx, row in enumerate(
    ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
        values_only=True
    ),
    start=start_row
):
    print(f"\nRow {idx}:")
    for i in range(len(headers)):
        print(f"  {headers[i]}: {row[i]},")
    
